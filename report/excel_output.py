from openpyxl import Workbook
from dotenv import load_dotenv
from copy import copy
import openpyxl
import random
import json
import os

load_dotenv()

MIN_ROW=int(os.getenv('MIN_ROW'))
MAX_ROW=int(os.getenv('MAX_ROW'))
MIN_COL=int(os.getenv('MIN_COL'))
MAX_COL=int(os.getenv('MAX_COL'))
_SEP = ":"

class ExcelReportOutput:
    def __init__(self, template_path, output_path, input_path):
        self.template_path = template_path
        self.output_path = output_path
        self.input_path = input_path
        print(self.template_path)
        print(self.output_path)
    
    def read_template(self) -> any:
        cells_dict = dict()
        try:
            template = openpyxl.load_workbook(self.template_path)
            print('template has been opened')
            ws = template.active
            for row in ws.iter_rows(min_row=MIN_ROW, max_row=MAX_ROW, min_col=MIN_COL, max_col=MAX_COL):
                for cell in row:
                    if cell.value is not None:         
                        cells_dict[cell.coordinate] = cell.value   
        except Exception as e: 
            print(e)
            print('failed to load file')
        return cells_dict

    def copy_template_format(self, ws):
        try:
            template = openpyxl.load_workbook(self.template_path)
            print('template has been opened')
            template_ws = template.active

            for row in template_ws.iter_rows(min_row=MIN_ROW, max_row=MAX_ROW, min_col=MIN_COL, max_col=MAX_COL):
                for cell in row:
                    self.copy_cell(cell, ws[cell.coordinate])
            # Copy column widths
            for col in range(MIN_COL, MAX_COL + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                template_width = template_ws.column_dimensions[col_letter].width
                if template_width:
                    ws.column_dimensions[col_letter].width = template_width

            # Copy row heights
            for row in range(MIN_ROW, MAX_ROW + 1):
                template_height = template_ws.row_dimensions[row].height
                if template_height:
                    ws.row_dimensions[row].height = template_height

        except Exception as e: 
            print(e)
            print('Failed to load file')

    def getValueFor(self, value):
        return self.switch(value)

    def copy_cell(self, source_cell, target_cell):
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

    def output(self):
        wb = Workbook()
        ws = wb.active

        print('copy the format')
        self.copy_template_format(ws)

        cells_dict = self.read_template()
        print(cells_dict)

        for cell in cells_dict:
            print(cell, _SEP, cells_dict[cell])
            val = self.getValueFor(cells_dict[cell])
            ws[cell] = val
            print(f"setting {cell}={val}")

        wb.save(self.output_path)

    def switch(self, value) -> any:
        print('value=>', value)
        match value:
            case "{o_t_1}":
                return self.extractFromJson(value)
            case "{o_t_2}":
                return self.extractFromJson(value)
            case "{o_t_3}":
                return self.extractFromJson(value)
            case "{o_t_4}":
                return self.extractFromJson(value)
            case "{o_t_5}":
                return self.computeResult()
            case _:
                return value

    def computeResult(self):
        return random.randint(200, 4000)

    def extractFromJson(self, value):
        f = open(self.input_path)
        data = json.load(f)
        client = data["clients"][0]
        if value == "{o_t_1}":
            return client["netDelivery"][0]["reach"]
        if value == "{o_t_2}":
            return client["netDelivery"][0]["reachPercent"]
        if value == "{o_t_3}":
            return client["netDelivery"][0]["frequency"]
        if value == "{o_t_4}":
            return client["netDelivery"][1]["reach"]
        if value == "{o_t_5}":
            return client["netDelivery"][2]["reachPercent"]
        if value == "{o_t_5}":
            return client["netDelivery"][3]["frequency"]
        return None
